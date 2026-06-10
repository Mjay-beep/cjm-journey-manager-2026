#!/usr/bin/env python3
"""CJM ↔ SBF ↔ BSS 매핑 초안 생성기 (Phase 1 PoC).

입력(읽기 전용):
  - sbf/1.0 ver/SKT_고객표준여정 관리 시트 (CX팀) - 고객여정 Manager.csv
  - sbf/SKT_Business_Framework_v2.0.xlsx  (0.Business Framework / 3.SSF(L3)-화면 맵핑)
  - sbf/SKT_Business_Framework_v1.9_작업중.csv  (업무ID → L3 ID 다리)

출력(sbf-cjm-mapping/ 안):
  - mapping_draft.csv     CX팀 검수용 관리 테이블
  - mapping_data.json     뷰어용 데이터
  - coverage_report.md    커버리지/갭 리포트

기존 시트·사이트·GAS는 일절 건드리지 않는다.
"""
import csv, json, re, os, sys
from collections import defaultdict

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(BASE, 'sbf-cjm-mapping')
CJM_CSV = os.path.join(BASE, 'sbf', '1.0 ver', 'SKT_고객표준여정 관리 시트 (CX팀) - 고객여정 Manager.csv')
SBF_XLSX = os.path.join(BASE, 'sbf', 'SKT_Business_Framework_v2.0.xlsx')
SBF_V19 = os.path.join(BASE, 'sbf', 'SKT_Business_Framework_v1.9_작업중.csv')

# ── 1. CJM 파싱 ──────────────────────────────────────────────
def parse_cjm():
    rows = list(csv.reader(open(CJM_CSV, encoding='utf-8')))
    nodes = []
    l1 = l2c = l2 = l3c = l3 = ''
    for r in rows:
        if len(r) < 9:
            continue
        c = [x.strip() for x in r[:9]]
        # 데이터 행 판별: L4 코드가 패턴을 가짐
        if not re.match(r'^[a-z]+_[a-z0-9_]+$', c[5] or ''):
            continue
        if c[0]: l1 = c[0].splitlines()[0].strip()
        if c[1]: l2c, l2 = c[1], c[2].replace('\n', ' ').strip()
        if c[3]: l3c, l3 = c[3], c[4].replace('\n', ' ').strip()
        nodes.append(dict(l1=l1, l2_code=l2c, l2=l2, l3_code=l3c, l3=l3,
                          l4_code=c[5], l4=c[6].replace('\n', ' ').strip(),
                          channel=c[7], note=c[8]))
    return nodes

# ── 2. SBF 파싱 ──────────────────────────────────────────────
# (B)범위: 고객접점 업무 포함 / 순수 백오피스 제외
EXCLUDE_1DEPTH = {
    '약관 거버넌스', '통합약관', '시스템 운영', '회계·세무 운영', '조직·사용자·운영',
    '멀티태넌시 운영', '모니터링·품질·리스크', '공지·알림 거버넌스',
    '자산/리소스 운영', '파트너 운영', '유통망 운영',
}

def parse_sbf():
    import openpyxl
    wb = openpyxl.load_workbook(SBF_XLSX, read_only=True, data_only=True)
    # v1.9: 업무ID → L3 ID들 (col1=업무ID, col13=L3 ID)
    biz2l3 = defaultdict(set)
    for r in list(csv.reader(open(SBF_V19, encoding='utf-8-sig')))[1:]:
        if len(r) > 13 and r[1]:
            for i in re.findall(r'BZ-[A-Z]{3}[A-Z0-9]+-\d+', r[13]):
                biz2l3[r[1].strip()].add(i)
    # 시트3: L3 ID → 화면들
    l3_screens = defaultdict(list)
    for r in wb['3. SSF(L3)-화면 맵핑'].iter_rows(min_row=2, values_only=True):
        key = str(r[0]) if r[0] else ''
        if key.startswith('BZ-') and r[2]:
            l3_screens[key].append({'pgm': str(r[2]).strip(), 'name': str(r[3] or '').strip()})
    # 메인시트
    tasks = {}
    for r in wb['0.Business Framework'].iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        bid = str(r[0]).strip()
        d1 = str(r[4] or '').strip()
        t = dict(biz_id=bid, gubun=str(r[3] or '').strip(), d1=d1,
                 d2=str(r[5] or '').strip(), d3=str(r[6] or '').strip(),
                 in_scope=d1 not in EXCLUDE_1DEPTH)
        screens = []
        for l3 in sorted(biz2l3.get(bid, [])):
            screens += l3_screens.get(l3, [])
        # 중복 화면 제거
        seen, uniq = set(), []
        for s in screens:
            if s['pgm'] not in seen:
                seen.add(s['pgm']); uniq.append(s)
        t['screens'] = uniq
        tasks[bid] = t
    return list(tasks.values())

# ── 3. 단계축 crosswalk: CJM L1 → 허용 SBF 1Depth ────────────
CROSSWALK = {
    '인입':        {'채널', 'CVM&CJM&C360', '상담&채널', '1.탐색', '고객 여정', 'CRM 운영'},
    '탐색':        {'1.탐색', '상품', '상담&채널', '채널', '고객 여정', 'CVM&CJM&C360'},
    '구매/가입/개통': {'3.계약·개통', '4.배송·설치', '2.회원·계정',
                    '가입/변경/해지(가입/변경/해지(Order))/풀필먼트 운영',
                    '영업/판매', '채널', '고객 여정', '상품'},
    '서비스 이용':   {'2.회원·계정', '5.이용·변경', '7.청구·수납', '수납·미납 운영',
                    '청구 운영', '8.채권·조정', '상담&채널', '채널', '고객 여정', 'CRM 운영'},
    '유지 및 확장':  {'5.이용·변경', 'CVM&CJM&C360', '영업/판매', '3.계약·개통',
                    '가입/변경/해지(가입/변경/해지(Order))/풀필먼트 운영', '고객 여정', '상품', '채널'},
    '이탈':        {'9.해지·잔존', '8.채권·조정', '수납·미납 운영', '5.이용·변경', '고객 여정', '상담&채널'},
}

# ── 4. 의미매칭: 키워드 + 동의어 ─────────────────────────────
# CJM 표현 ↔ SBF 표현이 다른 경우의 동의어 묶음 (같은 묶음 안이면 매칭 인정)
SYNONYMS = [
    {'미납', '연체', '체납', '추심', '채권'},
    {'납부', '수납', '결제', '자동이체', '카드납', '은행납'},
    {'요금', '청구', '청구서', '요금조회', '빌링'},
    {'해지', '이탈', '잔존', '해약'},
    {'가입', '개통', '청약', '신규'},
    {'기기변경', '기변', '단말변경'},
    {'번호이동', '번이', 'MNP'},
    {'휴대폰', '단말', '단말기', '기기', '스마트폰', '디바이스'},
    {'유심', 'USIM', 'eSIM', '이심'},
    {'인터넷', 'B tv', 'btv', 'IPTV', '유선'},
    {'부가서비스', '부가상품', '부가'},
    {'구독', 'T우주', '우주'},
    {'멤버십', '혜택', '쿠폰', '포인트', '제휴'},
    {'로밍', '해외', 'baro'},
    {'탐색', '조회', '검색', '찾기', '둘러보기'},
    {'추천', '개인화', '맞춤', '오퍼'},
    {'상담', '챗봇', '문의', 'CS', '고객센터', '콜센터', 'ARS'},
    {'분실', '습득', '보안', '정지', '일시정지'},
    {'명의변경', '명변', '양도'},
    {'번호변경', '번변'},
    {'배송', '풀필먼트', '택배', '수령'},
    {'예약', '사전예약'},
    {'캠페인', '광고', '프로모션', '이벤트', '마케팅'},
    {'푸시', '알림', '알림톡', 'PUSH', '메시지'},
    {'결합', '가족결합', '온가족'},
    {'요금제', '플랜', '요금상품'},
    {'할부', '약정', '위약금', '공시지원', '선택약정'},
    {'환불', '과오납', '반환'},
    {'회원', '계정', 'T ID', '아이디', '로그인', '인증'},
    {'데이터', '통화', '사용량', '실시간잔여'},
    {'매장', '대리점', '판매점', '오프라인', '직영'},
    {'정보변경', '개인정보', '고객정보'},
]
SYN_INDEX = {}
for group in SYNONYMS:
    for w in group:
        SYN_INDEX.setdefault(w.lower(), set()).update(x.lower() for x in group)

STOPWORDS = {'서비스', '관리', '조회', '신청', '변경', '확인', '이용', '처리', '진입',
             '및', '관련', '기타', '정보', '내역', '안내', '고객', '운영', '등록',
             '메뉴', '페이지', '화면', '목록', '상세', '버튼', '클릭', '선택', '입력'}

def extract_terms(text):
    """한글/영문 토큰 추출 (2자 이상, 불용어 제외)"""
    toks = re.findall(r'[가-힣]{2,}|[A-Za-z]{2,}', text or '')
    out = set()
    for t in toks:
        t = t.strip()
        if t and t not in STOPWORDS:
            out.add(t.lower())
            # 복합어 내부의 핵심어도 등록 (예: '미납처리' → '미납')
            for key in SYN_INDEX:
                if len(key) >= 2 and key in t.lower():
                    out.add(key)
    return out

def expand(terms):
    out = set(terms)
    for t in terms:
        out |= SYN_INDEX.get(t, set())
    return out

def match_score(cjm_terms_raw, sbf_terms_raw):
    """(점수, 근거단어목록). 직접 일치 > 동의어 일치."""
    direct = cjm_terms_raw & sbf_terms_raw
    syn = (expand(cjm_terms_raw) & sbf_terms_raw) - direct
    score = sum(min(len(w), 4) for w in direct) * 2 + sum(min(len(w), 4) for w in syn)
    return score, sorted(direct), sorted(syn)

# ── 4.5 수기 정밀검증 오버라이드 (미납 쇼케이스 등) ─────────────
# (CJM L3코드, SBF 업무ID): (신뢰도, 근거) — 자동 매칭 결과에 추가/승격
CURATED = {
    # 고객 "미납 요금 납부" ↔ 미납 상담·셀프채널·안내 업무
    ('use_pay_do', 'B8101'): ('상', '수기 정밀검증: 고객 미납 납부 ↔ 미납상담&채널관리 (고객미납내역조회·청구미납금액 화면)'),
    ('use_pay_do', 'B8103'): ('상', '수기 정밀검증: 고객 미납 셀프 처리 ↔ 미납Self채널'),
    ('use_pay_do', 'BC093'): ('상', '수기 정밀검증: 미납SELF채널 연동 (Self-PTP납부가능일기준관리 화면)'),
    ('use_pay_do', 'B8012'): ('상', '수기 정밀검증: 미납 안내 (미납알림톡 등 채널발송 화면 33개)'),
    # 고객 "미납 요금 조회" ↔ 미납 내역 조회 업무
    ('use_pay_hist', 'B8101'): ('상', '수기 정밀검증: 미납 요금 조회 ↔ 미납상담&채널관리 (고객미납내역조회 ZCOLSCNS02880)'),
    ('use_pay_hist', 'B8014'): ('중', '수기 정밀검증: 타사미납 내역 확인 (통신체납/미환급금 조회 화면)'),
}

# ── 5. 매핑 생성 ─────────────────────────────────────────────
def build():
    cjm = parse_cjm()
    sbf = parse_sbf()
    in_scope = [t for t in sbf if t['in_scope']]

    # CJM을 L3 단위로 그룹 (L4들은 매칭 텍스트로 흡수)
    groups = {}
    for n in cjm:
        g = groups.setdefault(n['l3_code'], dict(
            l1=n['l1'], l2_code=n['l2_code'], l2=n['l2'],
            l3_code=n['l3_code'], l3=n['l3'], l4s=[]))
        g['l4s'].append({'code': n['l4_code'], 'name': n['l4'], 'channel': n['channel']})

    results = []   # 매핑 행들
    for g in groups.values():
        allowed = CROSSWALK.get(g['l1'], set())
        l3_terms = extract_terms(g['l2'] + ' ' + g['l3'])
        cands = []
        for t in in_scope:
            stage_ok = t['d1'] in allowed
            sbf_terms = extract_terms(t['d2'] + ' ' + t['d3'])
            # L3 레벨 매칭
            sc, direct, syn = match_score(l3_terms, sbf_terms)
            best_l4 = ''
            # L4 레벨 보강: 개별 L4명이 강하게 맞으면 가산
            for l4 in g['l4s']:
                s4, d4, _ = match_score(extract_terms(l4['name']), sbf_terms)
                if s4 > sc:
                    sc, direct = s4, d4
                    best_l4 = l4['name']
            if sc <= 0:
                continue
            if not stage_ok:
                sc *= 0.4   # 단계축 밖이면 감점 (완전 배제는 안 함 — 운영성 업무 대비)
            cands.append((sc, t, direct, syn, best_l4, stage_ok))
        cands.sort(key=lambda x: -x[0])
        for sc, t, direct, syn, best_l4, stage_ok in cands[:6]:  # 노드당 상위 6건
            if sc >= 12 and stage_ok and direct:
                conf = '상'
            elif sc >= 6 and (stage_ok or direct):
                conf = '중'
            elif sc >= 4:
                conf = '하'
            else:
                continue
            why = []
            if direct: why.append('용어일치: ' + '·'.join(direct[:4]))
            if syn: why.append('동의어: ' + '·'.join(syn[:3]))
            why.append('단계축 ' + ('일치' if stage_ok else '불일치'))
            if best_l4: why.append(f'L4 [{best_l4}] 기준')
            results.append(dict(
                l1=g['l1'], l2=g['l2'], l3=g['l3'], l3_code=g['l3_code'],
                l4_hint=best_l4, biz_id=t['biz_id'], sbf_d1=t['d1'], sbf_d2=t['d2'],
                sbf_d3=t['d3'], screens=t['screens'], conf=conf, score=round(sc, 1),
                why=' / '.join(why)))

    # 수기 오버라이드 적용: 기존 행 승격 또는 신규 삽입
    by_task = {t['biz_id']: t for t in sbf}
    existing = {(r['l3_code'], r['biz_id']): r for r in results}
    for (l3c, bid), (conf, why) in CURATED.items():
        if l3c not in groups or bid not in by_task:
            print(f'경고: CURATED 키 불일치 {l3c}/{bid}', file=sys.stderr)
            continue
        if (l3c, bid) in existing:
            existing[(l3c, bid)].update(conf=conf, why=why)
        else:
            g, t = groups[l3c], by_task[bid]
            results.append(dict(
                l1=g['l1'], l2=g['l2'], l3=g['l3'], l3_code=l3c, l4_hint='',
                biz_id=bid, sbf_d1=t['d1'], sbf_d2=t['d2'], sbf_d3=t['d3'],
                screens=t['screens'], conf=conf, score=99, why=why))
    return cjm, sbf, groups, results

def write_outputs(cjm, sbf, groups, results):
    os.makedirs(OUT, exist_ok=True)
    # 1) 검수용 CSV
    with open(os.path.join(OUT, 'mapping_draft.csv'), 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f)
        w.writerow(['CJM_L1', 'CJM_L2', 'CJM_L3', 'CJM_L3코드', '관련_L4(근거)',
                    'SBF_업무ID', 'SBF_1Depth', 'SBF_2Depth', 'SBF_3Depth(업무명)',
                    'BSS_화면수', 'BSS_대표화면(최대5)', '신뢰도', '매칭근거',
                    '검수상태(승인/수정/반려)', '검수메모'])
        for r in sorted(results, key=lambda x: (x['l1'], x['l2'], x['l3'], {'상': 0, '중': 1, '하': 2}[x['conf']])):
            top = ' | '.join(f"{s['pgm']} {s['name']}" for s in r['screens'][:5])
            w.writerow([r['l1'], r['l2'], r['l3'], r['l3_code'], r['l4_hint'],
                        r['biz_id'], r['sbf_d1'], r['sbf_d2'], r['sbf_d3'],
                        len(r['screens']), top, r['conf'], r['why'], '', ''])
    # 2) 뷰어용 JSON
    tree = []
    by_l1 = defaultdict(lambda: defaultdict(list))
    for g in groups.values():
        by_l1[g['l1']][g['l2_code'] + '|' + g['l2']].append(g)
    L1_ORDER = ['인입', '탐색', '구매/가입/개통', '서비스 이용', '유지 및 확장', '이탈']
    res_by_l3 = defaultdict(list)
    for r in results:
        res_by_l3[r['l3_code']].append(r)
    for l1 in L1_ORDER:
        l2s = []
        for key, gs in by_l1.get(l1, {}).items():
            _, l2name = key.split('|', 1)
            l3s = []
            for g in gs:
                maps = [dict(biz_id=m['biz_id'], d1=m['sbf_d1'], d2=m['sbf_d2'], d3=m['sbf_d3'],
                             conf=m['conf'], why=m['why'],
                             screens=m['screens'][:30]) for m in
                        sorted(res_by_l3.get(g['l3_code'], []), key=lambda x: {'상': 0, '중': 1, '하': 2}[x['conf']])]
                l3s.append(dict(code=g['l3_code'], name=g['l3'],
                                l4s=[x['name'] for x in g['l4s']], maps=maps))
            l2s.append(dict(name=l2name, l3s=l3s))
        tree.append(dict(l1=l1, l2s=l2s))
    with open(os.path.join(OUT, 'mapping_data.json'), 'w', encoding='utf-8') as f:
        json.dump(tree, f, ensure_ascii=False, indent=1)
    # 2.5) 뷰어 HTML (JSON 임베드 — file:// 로 바로 열림)
    tpl = open(os.path.join(OUT, 'viewer_template.html'), encoding='utf-8').read()
    with open(os.path.join(OUT, 'mapping_viewer.html'), 'w', encoding='utf-8') as f:
        f.write(tpl.replace('__DATA__', json.dumps(tree, ensure_ascii=False)))
    # 3) 커버리지 리포트
    mapped_l3 = set(res_by_l3)
    all_l3 = set(groups)
    unmapped = [groups[c] for c in sorted(all_l3 - mapped_l3)]
    conf_cnt = defaultdict(int)
    for r in results:
        conf_cnt[r['conf']] += 1
    in_scope = [t for t in sbf if t['in_scope']]
    used_biz = set(r['biz_id'] for r in results)
    unused = [t for t in in_scope if t['biz_id'] not in used_biz]
    with open(os.path.join(OUT, 'coverage_report.md'), 'w', encoding='utf-8') as f:
        f.write('# CJM ↔ SBF 매핑 커버리지 리포트 (자동생성 초안)\n\n')
        f.write(f'- CJM L3 노드: {len(all_l3)}개 중 **{len(mapped_l3)}개 매핑** ({len(mapped_l3)/len(all_l3)*100:.0f}%)\n')
        f.write(f'- 매핑 연결 수: {len(results)}건 (상 {conf_cnt["상"]} / 중 {conf_cnt["중"]} / 하 {conf_cnt["하"]})\n')
        f.write(f'- SBF 대상범위(고객접점) 업무: {len(in_scope)}개 중 {len(used_biz)}개 사용\n')
        f.write(f'- BSS 화면 보유 연결: {sum(1 for r in results if r["screens"])}건\n\n')
        f.write('## 매핑 안 된 CJM L3 (갭 — 임원 논의 대상)\n\n')
        for g in unmapped:
            f.write(f'- [{g["l1"]} > {g["l2"]}] {g["l3"]} (`{g["l3_code"]}`)\n')
        f.write('\n## CJM에 연결되지 않은 SBF 고객접점 업무 (역방향 갭)\n\n')
        for t in sorted(unused, key=lambda x: x['d1']):
            f.write(f'- [{t["d1"]} > {t["d2"]}] {t["d3"]} (`{t["biz_id"]}`)\n')
    print(f'CJM L3 {len(all_l3)} → 매핑 {len(mapped_l3)} ({len(mapped_l3)/len(all_l3)*100:.0f}%) | '
          f'연결 {len(results)}건 (상{conf_cnt["상"]}/중{conf_cnt["중"]}/하{conf_cnt["하"]}) | '
          f'SBF사용 {len(used_biz)}/{len(in_scope)}')

if __name__ == '__main__':
    cjm, sbf, groups, results = build()
    write_outputs(cjm, sbf, groups, results)
